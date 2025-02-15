from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.drawing.image import Image


from plantapp.models import Plant
fontBold = Font(bold=True,name='Calibri',)
fontGenus= Font(bold=True,size=18,name='Calibri')
center = Alignment(horizontal='center',vertical='center')
centerWrap = Alignment(horizontal='center',vertical='center',wrap_text=True)
class PriceRender:
    template_name = "template.xls"
    dest_filename="text.xlsx"
    columnt_width = {
        'A':2,
        'B':20,
        'C':20,
        'D':6,
        'E':6,
        'F':6,
        'G':6,
        'H':16,
    }
    def __init__(self,queryset,title=''):
        self.wb= Workbook()
        self.queryset=queryset
        self.ws = self.wb.active
        self.title= title
    def set_layout(self):
        for k,v in self.columnt_width.items():
            self.ws.column_dimensions[k].width =v
        for i in range(1,6):
            self.ws.merge_cells(f'C{i}:H{i}')
        for i in range(6,9):
            self.ws.merge_cells(f'B{i}:H{i}')

    def head(self):
        self.ws['C1'] = 'Питомник садовых растений «Красная Гвоздика»'
        self.ws['C1'].font =fontBold
        self.ws['C2'] = 'ИП Черепанова Вера Ивановна'
        self.ws['C2'].font = fontBold

        self.ws['C3'] = 'Пермский край, Пермский район, Култаевское с/п, Квартал "Красная гвоздика", ул. Цветочная, 2 '
        self.ws['C4'] = 'krasnaya-gvozdika@yandex.ru '
        self.ws['C5'] = 'тел. 8-922-640-66-83, 8-922-640-86-04, 8-922-640-66-49'
        self.ws['C6'] = ''

        self.ws['B7'] = 'Прайс'
        self.ws['B8'] = self.title

        for i in range(1,6):
            self.ws[f'C{i}'].alignment = center

        for i in range(6,8):
            self.ws[f'B{i}'].alignment = center
        self.ws['B9'] = 'Русское название'
        self.ws['C9'] = 'Латинское название'
        self.ws['D9'] = 'Диаметр ствола на высоте 1 м. (см)'
        self.ws['E9'] = 'Объем контейнера, Л'
        self.ws['F9'] = 'Высота, см'
        self.ws['G9'] = 'Цена, руб.'
        self.ws['H9'] = 'Примечание'
        img = Image('logo.jpeg')
        img.anchor = 'B1'
        self.ws.add_image(img)
        for i in 'BCDEFGH':
            self.ws[f'{i}9'].font = fontBold
    def render_item(self,item,row):
        i = row
        for g in item.goods_set.all():
            i+=1
            self.ws[f'D{i}'] =''
            self.ws[f'F{i}'] = g.pot_size()
            self.ws[f'G{i}'] =g.price
        for r in 'BC':
            self.ws.merge_cells(f'{r}{row}:{r}{i}')

        self.ws.merge_cells(f'D{row}:H{row}')
        self.ws[f'B{row}'] = item.species_name() + '. '+item.sort
        self.ws[f'C{row}'] = item.species_name() +'. '+ item.sort_latin
        description = item.get_description()
        self.ws[f'D{row}'] = description
        self.ws[f'D{row}'].alignment= centerWrap
        self.ws.row_dimensions[row].height = (len(description)//50+1)*20

        self.ws[f'B{row}'].alignment = centerWrap
        self.ws[f'C{row}'].alignment = centerWrap

        self.ws[f'B{row}'].font = fontBold
        self.ws[f'C{row}'].font = fontBold


        return i
    def render_genus(self,genus,row):
        print(genus)
        self.ws.merge_cells(f'B{row}:C{row}')
        self.ws[f'B{row}'] = genus.name
        self.ws[f'B{row}'].alignment = centerWrap
        self.ws[f'B{row}'].font = fontGenus
        self.ws.row_dimensions[row].height =30

        return row+1

    def get_queryset(self):
        return self.queryset
    def body(self):
        row=10
        genus = None
        for item in self.get_queryset():
            row+=1
            if item.species.genus !=genus:
                genus = item.species.genus
                row = self.render_genus(genus,row)
            row = self.render_item(item,row)
        return row
    def set_border(self,r,border):
        for r in self.ws[r]:
            for c in r:
                c.border = border

    def border(self,row):
        thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
        right =  Border(
                     right=Side(style='medium'),
                     )

        left=  Border(left=Side(style='medium'),
                     )

        top =  Border(
                     top=Side(style='medium'),
        )

        bottom =  Border(
                     bottom=Side(style='medium')
        )


        self.set_border(f'B9:H{row}',thin_border)
        self.set_border(f'B9:B{row}',left)
        self.set_border(f'H9:H{row}',right)
        self.set_border(f'B{row}:H{row}',bottom)
        self.set_border('B9:H9',top)
        self.set_border('B10:H10',top)

    def render(self):
        self.head()
        self.set_layout()
        row = self.body()
        self.border(row)

    def save(self):
        self.wb.save(filename = self.dest_filename)

    def get_steam(self):
        self.render()
        return save_virtual_workbook(self.wb)
